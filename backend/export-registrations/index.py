import json
import os
import io
import boto3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import base64


S3_KEY = 'registrations/data.json'


def get_s3():
    return boto3.client(
        's3',
        endpoint_url='https://bucket.poehali.dev',
        aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
        aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY']
    )


def load_registrations(s3):
    try:
        obj = s3.get_object(Bucket='files', Key=S3_KEY)
        return json.loads(obj['Body'].read().decode('utf-8'))
    except BaseException:
        return []


def handler(event: dict, context) -> dict:
    """Выгружает список зарегистрированных участников в Excel-файл."""
    if event.get('httpMethod') == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Max-Age': '86400'
            },
            'body': ''
        }

    s3 = get_s3()
    registrations = load_registrations(s3)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Регистрации'

    headers = ['#', 'Фамилия Имя', 'Телефон', 'Email для счёта', 'Компания', 'ИНН', 'КПП', 'Юр. адрес', 'Мест', 'Дата заявки']
    header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    col_widths = [5, 25, 20, 30, 30, 15, 12, 40, 8, 20]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 20

    fields = ['id', 'fullName', 'phone', 'billingEmail', 'company', 'inn', 'kpp', 'legalAddress', 'seats', 'createdAt']
    for row_num, reg in enumerate(registrations, 2):
        for col_num, field in enumerate(fields, 1):
            cell = ws.cell(row=row_num, column=col_num, value=reg.get(field, ''))
            cell.alignment = Alignment(vertical='center', wrap_text=(col_num == 8))
        if row_num % 2 == 0:
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).fill = PatternFill(
                    start_color='F0F4F8', end_color='F0F4F8', fill_type='solid'
                )

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    excel_b64 = base64.b64encode(buffer.read()).decode('utf-8')

    return {
        'statusCode': 200,
        'headers': {
            'Access-Control-Allow-Origin': '*',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': 'attachment; filename="registrations.xlsx"',
        },
        'body': excel_b64,
        'isBase64Encoded': True
    }